import random
import openpyxl
from aiogram import Router, F
from aiogram.types import Message, CallbackQuery
from aiogram.filters import CommandStart, Command, or_f
from aiogram.fsm.state import StatesGroup, State
from aiogram.fsm.context import FSMContext
from aiogram.types import ReplyKeyboardRemove
import app.keyboards as kb
from config import ADMIN_ID, questions_count

router = Router()

class Test(StatesGroup):
    count = State()
    play = State()
    feedback = State()
    question_index = State()


@router.message(CommandStart())
async def start(message: Message):
    questions_count.setdefault(message.from_user.id, 15)
    await message.answer(text = f'Привет <b>{message.from_user.first_name}</b>,\nЭто игра в которой будет тест и вы должны ответить на вопросы!', reply_markup = kb.start, parse_mode = 'html')


@router.message(Command('admin'))
async def admin(message: Message):
    if message.from_user.id == ADMIN_ID:
        await message.answer(text = f'Здравствуйте, {message.from_user.full_name}')
        await message.answer(text = 'Добро Пожаловать в Админ Панель!')
    else: 
        print(message.from_user.id)
        await message.answer('Нет такой команды!')
    

@router.message(or_f((F.text.lower() == 'помощь') | (F.text.lower() == 'о игре'), Command('help')))
async def help(message: Message):
    await message.answer(text = 'В этой игре ты должен будешь проходить тест отвечая на вопросы с 4 вариантами ответа. В конце игра покажет результаты!', reply_markup = kb.back)


@router.message(or_f((F.text.lower() == 'настройки'), Command('settings')))
async def settings(message: Message):
    await message.answer(text = f'Сейчас {questions_count[message.from_user.id]} вопросов, минимальное количество 5 вопросов и максимальное количество 50 вопросов', reply_markup = kb.settings)


@router.message(or_f((F.text.lower() == 'играть'), Command('play')))
async def play(message: Message, state: FSMContext):
    await state.set_state(Test.count)
    await message.answer(text='Введите номер теста (от 1 до 50)', reply_markup = kb.back)


@router.message(Test.count)
async def play_1(message: Message, state: FSMContext):
    try:
        count = int(message.text)
        if 1 <= count <= 50:
            await state.set_state(Test.play)
            wb = openpyxl.load_workbook('test.xlsx')
            test = wb[f'Test-{count}']
            all_rows = list(test.iter_rows(max_row = 49, values_only = True))
            try:
                random_rows = random.sample(all_rows, questions_count[message.from_user.id] - 1)
            except KeyError:
                questions_count[message.from_user.id] = 15
                random_rows = random.sample(all_rows, questions_count[message.from_user.id] - 1)
            question = random_rows[0]
            feedback = list(test.iter_rows(min_row=test.max_row, max_row=test.max_row, values_only=True))[0]
            await message.answer(text = 'Игра началась')
            keyboard = kb.test(question)
            await message.answer(text = f'Вопрос {1}: {question[0]}', reply_markup = keyboard)
            await state.update_data(play = random_rows)
            await state.update_data(feedback = feedback)
            await state.update_data(count = 0)
            await state.update_data(question_index = 0)
        else:
            await message.answer(text = 'Номер теста должен быть от 1 до 50')
    except ValueError:
        await message.answer(text = 'Некорректный ввод. Пожалуйста, введите число')

@router.message(Test.play)
async def play_2(message: Message, state: FSMContext):
    data = await state.get_data()
    question_index = data['question_index']
    count = data['count']
    random_rows = data['play']
    last_row = data['feedback']
    try:
        question = random_rows[question_index]
    except IndexError:
        question = last_row
    user_answer = message.text
    answer = question[5]
    if str(user_answer).lower() == str(answer).lower():
        await message.answer(text = 'Верно!', reply_markup = kb.end)
        count += 1
        await state.update_data(count = count)
    else:
        await message.answer(text = f'Неверно! Правильный ответ: {answer}', reply_markup = kb.end)
    
    if question_index < questions_count[message.from_user.id] - 2:
        next_question = random_rows[question_index + 1]
        keyboard = kb.test(next_question)
        await message.answer(text = f'Вопрос {question_index + 2}\n{next_question[0]}', reply_markup = keyboard)
    elif question_index == questions_count[message.from_user.id] - 2:
        keyboard = kb.test(last_row)
        await message.answer(text = f'Вопрос {question_index + 2}\n{last_row[0]}', reply_markup = keyboard)
    else:
        await message.answer(text = f'Тест завершен, Вы набрали {count}/{questions_count[message.from_user.id]}.\nСпасибо за игру!', reply_markup = kb.back)
        await message.answer(text = f'Сыграть ещё раз? /play')
        await state.clear()
    question_index += 1
    await state.update_data(question_index = question_index)


@router.callback_query(F.data == 'again')
async def back(callback: CallbackQuery, state: FSMContext):
    await state.clear()
    await callback.answer()
    await state.set_state(Test.count)
    await callback.message.answer(text='Введите номер теста (от 1 до 50)', reply_markup = kb.back)


@router.callback_query(F.data == 'end')
async def back(callback: CallbackQuery, state: FSMContext):
    data = await state.get_data()
    question_index = data['question_index']
    count = data['count']
    await callback.answer()
    await callback.message.answer(text = f'Тест завершен, Вы набрали {count}/{question_index + 1}.\nСпасибо за игру!', reply_markup = kb.back)
    await callback.message.answer(text = f'Сыграть ещё раз? /play')
    await state.clear()


@router.callback_query(F.data == 'back')
async def back(callback: CallbackQuery):
    await callback.answer()
    await callback.message.answer(text = f'Привет <b>{callback.from_user.first_name}</b>,\nЭто игра в которой будет тест и вы должны ответить на вопросы!', reply_markup = kb.start, parse_mode = 'html')


@router.callback_query(F.data == '+')
async def plus(callback: CallbackQuery):
    if questions_count[callback.from_user.id] == 50:
        await callback.answer(text='Достигнуто максимальное количество вопросов', show_alert=True)
    else:
        questions_count[callback.from_user.id] += 5
        await callback.answer()
        await callback.message.edit_text(text=f'Сейчас {questions_count[callback.from_user.id]} вопросов, минимальное количество 5 вопросов и максимальное количество 50 вопросов', reply_markup=kb.settings)


@router.callback_query(F.data == '-')
async def minus(callback: CallbackQuery):
    if questions_count[callback.from_user.id] == 5:
        await callback.answer(text='Достигнуто минимальное количество вопросов', show_alert=True)
    else:
        questions_count[callback.from_user.id] -= 5
        await callback.answer()
        await callback.message.edit_text(text=f'Сейчас {questions_count[callback.from_user.id]} вопросов, минимальное количество 5 вопросов и максимальное количество 50 вопросов', reply_markup=kb.settings)
